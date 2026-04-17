from __future__ import annotations

import json
import re
from io import BytesIO

import sqlparse
from openpyxl import load_workbook

from app.schemas.dto import JsonFieldInput
from app.services.normalization import normalize_comment

HEADER_ALIASES = {
    "table_name": {"table_name", "table", "tablename", "表名", "数据表名", "表"},
    "column_name": {"column_name", "field_name", "column", "字段名", "列名", "字段英文名", "英文名"},
    "column_comment_zh": {
        "column_comment_zh",
        "comment_zh",
        "column_comment",
        "comment",
        "字段注释",
        "中文注释",
        "注释",
        "字段中文名",
    },
    "data_type": {"data_type", "datatype", "type", "字段类型", "数据类型", "类型"},
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-]+", "", text)
    return text


def _normalize_imported_column_name(value: object) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[\s\-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text.lower()


def _contains_chinese(value: object) -> bool:
    return bool(re.search(r"[\u4e00-\u9fff]", str(value or "")))


def _looks_like_identifier(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return bool(re.fullmatch(r"[A-Za-z_][A-Za-z0-9_\-\s]*", text))


def _resolve_header_mapping(header_row: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [_normalize_header(cell) for cell in header_row]
    for logical_name, aliases in HEADER_ALIASES.items():
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        for index, header_name in enumerate(normalized):
            if header_name in normalized_aliases:
                mapping[logical_name] = index
                break
    return mapping


def _parse_excel_mapping_template(rows: list[tuple[object, ...]], default_table_name: str) -> list[dict]:
    if not rows:
        return []

    max_columns = max((len(row) for row in rows), default=0)
    best_pair: tuple[int, int] | None = None
    best_score = 0

    for left_index in range(max_columns):
        for right_index in range(left_index + 1, max_columns):
            score = 0
            for row in rows[:100]:
                left = row[left_index] if left_index < len(row) else None
                right = row[right_index] if right_index < len(row) else None
                if _looks_like_identifier(left) and _contains_chinese(right):
                    score += 1
            if score > best_score:
                best_score = score
                best_pair = (left_index, right_index)

    if best_pair is None or best_score < 2:
        return []

    left_index, right_index = best_pair
    fields: list[dict] = []
    for row in rows:
        left = row[left_index] if left_index < len(row) else None
        right = row[right_index] if right_index < len(row) else None
        if not (_looks_like_identifier(left) and _contains_chinese(right)):
            continue
        column_name = _normalize_imported_column_name(left)
        comment = str(right).strip()
        fields.append(
            {
                "table_name": default_table_name,
                "column_name": column_name,
                "column_comment_zh": comment,
                "canonical_comment_zh": normalize_comment(comment),
                "data_type": None,
            }
        )
    return fields


def parse_sql_fields(sql_text: str) -> list[dict]:
    table_name = "unknown_table"
    parsed_fields: list[dict] = []
    comment_map: dict[tuple[str, str], str] = {}
    fallback_comment_map: dict[str, str] = {}

    for table_name_hint, column_name_hint, comment in re.findall(
        r"comment\s+on\s+column\s+`?([a-zA-Z0-9_]+)`?\.`?([a-zA-Z0-9_]+)`?\s+is\s+'([^']*)'",
        sql_text,
        flags=re.IGNORECASE,
    ):
        comment_map[(table_name_hint.lower(), column_name_hint.lower())] = comment
        fallback_comment_map[column_name_hint.lower()] = comment

    statements = sqlparse.parse(sql_text)
    for statement in statements:
        text = str(statement)
        table_match = re.search(r"create\s+table\s+`?([a-zA-Z0-9_]+)`?", text, flags=re.IGNORECASE)
        if table_match:
            table_name = table_match.group(1)
        for line in text.splitlines():
            cleaned = line.strip().rstrip(",")
            if not cleaned or cleaned.lower().startswith(("create table", "primary key", "key ", "unique ", "comment on ", ")")):
                continue
            column_match = re.match(
                r"`?([a-zA-Z0-9_]+)`?\s+([a-zA-Z0-9()_, ]+?)(?:\s+comment\s+'([^']+)')?$",
                cleaned,
                flags=re.IGNORECASE,
            )
            if not column_match:
                continue
            column_name, data_type, comment = column_match.groups()
            if not comment:
                comment = comment_map.get((table_name.lower(), column_name.lower()))
            if not comment:
                comment = fallback_comment_map.get(column_name.lower())
            parsed_fields.append(
                {
                    "table_name": table_name,
                    "column_name": column_name,
                    "column_comment_zh": comment,
                    "canonical_comment_zh": normalize_comment(comment) if comment else None,
                    "data_type": data_type.strip(),
                }
            )
    return parsed_fields


def parse_json_fields(items: list[JsonFieldInput]) -> list[dict]:
    fields: list[dict] = []
    for item in items:
        fields.append(
            {
                "table_name": item.table_name,
                "column_name": item.column_name,
                "column_comment_zh": item.column_comment_zh,
                "canonical_comment_zh": normalize_comment(item.column_comment_zh) if item.column_comment_zh else None,
                "data_type": item.data_type,
            }
        )
    return fields


def parse_excel_fields(file_bytes: bytes) -> list[dict]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    fields: list[dict] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = next((index for index, row in enumerate(rows) if any(cell not in (None, "") for cell in row)), None)
        if header_index is None:
            continue
        header_row = list(rows[header_index])
        header_mapping = _resolve_header_mapping(header_row)

        default_table_name = worksheet.title.strip() or "unknown_table"
        if "column_name" not in header_mapping:
            fields.extend(_parse_excel_mapping_template(rows, default_table_name))
            continue

        for row in rows[header_index + 1 :]:
            row_values = list(row)
            column_name_index = header_mapping["column_name"]
            column_name = row_values[column_name_index] if column_name_index < len(row_values) else None
            if column_name in (None, ""):
                continue

            table_name_index = header_mapping.get("table_name")
            comment_index = header_mapping.get("column_comment_zh")
            data_type_index = header_mapping.get("data_type")

            table_name = (
                row_values[table_name_index]
                if table_name_index is not None and table_name_index < len(row_values) and row_values[table_name_index] not in (None, "")
                else default_table_name
            )
            comment = (
                row_values[comment_index]
                if comment_index is not None and comment_index < len(row_values) and row_values[comment_index] not in (None, "")
                else None
            )
            data_type = (
                row_values[data_type_index]
                if data_type_index is not None and data_type_index < len(row_values) and row_values[data_type_index] not in (None, "")
                else None
            )

            fields.append(
                {
                    "table_name": str(table_name).strip(),
                    "column_name": _normalize_imported_column_name(column_name),
                    "column_comment_zh": str(comment).strip() if comment is not None else None,
                    "canonical_comment_zh": normalize_comment(str(comment)) if comment is not None else None,
                    "data_type": str(data_type).strip() if data_type is not None else None,
                }
            )

    return fields


def parse_txt_fields(text: str) -> list[dict]:
    stripped = text.strip()
    if not stripped:
        return []

    # JSON array or object-per-line
    if stripped.startswith("["):
        payload = json.loads(stripped)
        items = [JsonFieldInput(**item) for item in payload]
        return parse_json_fields(items)

    if stripped.startswith("{"):
        items = [JsonFieldInput(**json.loads(line)) for line in stripped.splitlines() if line.strip()]
        return parse_json_fields(items)

    # TSV-like rows: table_name \t column_name \t comment \t data_type
    fields: list[dict] = []
    for line in stripped.splitlines():
        raw = line.strip()
        if not raw:
            continue
        if "|" in raw:
            parts = [part.strip() for part in raw.split("|")]
        else:
            parts = [part.strip() for part in raw.split("\t")]
        if len(parts) < 2:
            continue
        table_name = parts[0] or "unknown_table"
        column_name = parts[1]
        comment = parts[2] if len(parts) > 2 and parts[2] else None
        data_type = parts[3] if len(parts) > 3 and parts[3] else None
        fields.append(
            {
                "table_name": table_name,
                "column_name": _normalize_imported_column_name(column_name),
                "column_comment_zh": comment,
                "canonical_comment_zh": normalize_comment(comment) if comment else None,
                "data_type": data_type,
            }
        )
    if fields:
        return fields

    for line in stripped.splitlines():
        raw = line.strip()
        if not raw:
            continue
        parts = raw.split()
        if len(parts) < 2:
            continue

        data_type = None
        if re.fullmatch(r"[A-Za-z]+(?:\(\d+(?:,\d+)?\))?", parts[-1]):
            data_type = parts[-1]
            parts = parts[:-1]
        if len(parts) < 2:
            continue

        if len(parts) >= 3 and _looks_like_identifier(parts[0]) and _looks_like_identifier(parts[1]) and _contains_chinese(" ".join(parts[2:])):
            table_name = parts[0]
            column_name = parts[1]
            comment = " ".join(parts[2:])
        elif _looks_like_identifier(parts[0]) and _contains_chinese(" ".join(parts[1:])):
            table_name = "unknown_table"
            column_name = parts[0]
            comment = " ".join(parts[1:])
        else:
            continue

        fields.append(
            {
                "table_name": table_name,
                "column_name": _normalize_imported_column_name(column_name),
                "column_comment_zh": comment,
                "canonical_comment_zh": normalize_comment(comment),
                "data_type": data_type,
            }
        )
    return fields
