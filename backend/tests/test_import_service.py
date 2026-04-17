from io import BytesIO

from openpyxl import Workbook

from app.schemas.dto import JsonFieldInput
from app.services.import_service import parse_excel_fields, parse_json_fields, parse_sql_fields, parse_txt_fields


def test_parse_sql_fields_extracts_comments():
    sql = """
    CREATE TABLE user_profile (
      user_id bigint COMMENT '用户编号',
      mobile_no varchar(32) COMMENT '联系电话'
    );
    """
    fields = parse_sql_fields(sql)
    assert len(fields) == 2
    assert fields[0]["column_comment_zh"] == "用户编号"


def test_parse_json_fields_extracts_items():
    items = parse_json_fields(
        [
            JsonFieldInput(
                table_name="user_profile",
                column_name="user_id",
                column_comment_zh="用户编号",
                data_type="bigint",
            )
        ]
    )
    assert items[0]["canonical_comment_zh"] is not None


def test_parse_sql_fields_supports_comment_on_column_syntax():
    sql = """
    create table t_gdfa
    (
        parcel_no varchar(255),
        parcel_name varchar(255)
    );

    comment on table t_gdfa is '供地方案';
    comment on column t_gdfa.parcel_no is '地块编号';
    comment on column t_gdfa.parcel_name is '地块名称';
    """
    fields = parse_sql_fields(sql)
    assert len(fields) == 2
    assert fields[0]["column_comment_zh"] == "地块编号"
    assert fields[1]["column_comment_zh"] == "地块名称"


def test_parse_excel_fields_supports_english_headers():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "user_profile"
    worksheet.append(["table_name", "column_name", "column_comment_zh", "data_type"])
    worksheet.append(["user_profile", "user_id", "用户编号", "bigint"])
    worksheet.append(["user_profile", "mobile_no", "联系电话", "varchar(64)"])

    buffer = BytesIO()
    workbook.save(buffer)

    fields = parse_excel_fields(buffer.getvalue())
    assert len(fields) == 2
    assert fields[0]["table_name"] == "user_profile"
    assert fields[0]["column_name"] == "user_id"
    assert fields[0]["column_comment_zh"] == "用户编号"
    assert fields[0]["canonical_comment_zh"] is not None


def test_parse_excel_fields_uses_sheet_name_when_table_name_missing():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "land_project"
    worksheet.append(["字段名", "中文注释", "数据类型"])
    worksheet.append(["parcel_no", "地块编号", "varchar(255)"])

    buffer = BytesIO()
    workbook.save(buffer)

    fields = parse_excel_fields(buffer.getvalue())
    assert len(fields) == 1
    assert fields[0]["table_name"] == "land_project"
    assert fields[0]["column_name"] == "parcel_no"


def test_parse_excel_fields_supports_left_english_right_chinese_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "land_direct_supply"
    worksheet.append(["BLOCK_NOTICE_NO", "子地块公告号"])
    worksheet.append(["parcel number", "宗地编号"])
    worksheet.append(["BLOCK_NAME", "地块名称"])

    buffer = BytesIO()
    workbook.save(buffer)

    fields = parse_excel_fields(buffer.getvalue())
    assert len(fields) == 3
    assert fields[0]["table_name"] == "land_direct_supply"
    assert fields[0]["column_name"] == "block_notice_no"
    assert fields[1]["column_name"] == "parcel_number"
    assert fields[2]["column_comment_zh"] == "地块名称"


def test_parse_txt_fields_supports_json_array():
    content = """
    [
      {"table_name":"user_profile","column_name":"user_id","column_comment_zh":"用户编号","data_type":"bigint"}
    ]
    """
    fields = parse_txt_fields(content)
    assert len(fields) == 1
    assert fields[0]["column_name"] == "user_id"
    assert fields[0]["canonical_comment_zh"] is not None


def test_parse_txt_fields_supports_tsv_rows():
    content = "user_profile\tmobile_no\t联系电话\tvarchar(64)\nuser_profile\tuser_id\t用户编号\tbigint"
    fields = parse_txt_fields(content)
    assert len(fields) == 2
    assert fields[0]["table_name"] == "user_profile"
    assert fields[0]["column_name"] == "mobile_no"


def test_parse_txt_fields_supports_pipe_delimited_rows():
    content = "user_profile|mobile_no|联系电话|varchar(64)\nuser_profile|user_id|用户编号|bigint"
    fields = parse_txt_fields(content)
    assert len(fields) == 2
    assert fields[0]["table_name"] == "user_profile"
    assert fields[0]["column_name"] == "mobile_no"


def test_parse_txt_fields_supports_space_separated_short_form():
    content = "parcel_no 地块编号 varchar(255)\nblock_notice_no 子地块公告号"
    fields = parse_txt_fields(content)
    assert len(fields) == 2
    assert fields[0]["column_name"] == "parcel_no"
    assert fields[0]["column_comment_zh"] == "地块编号"
    assert fields[0]["data_type"] == "varchar(255)"
